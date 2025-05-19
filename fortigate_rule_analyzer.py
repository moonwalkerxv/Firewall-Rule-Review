tool_info="""

Tool Name   : Fortigate Firewall Rule Analyzer v1.0
Version     : 1.0

Auther: Amol A. Pandhare
Email : pandhareamol96@gmail.com

* This application was created for myslef for security review of fortifate firewall rules.
* The information about the groups listed in the source, destination, and services columns is not included in the conventional firewall rule export. 
* To free up time for the real evaluation, we have attempted to automate repetitive and tedious tasks.
* The report will be a Microsoft Excel file in the.xlsx format with some basic formatting. Feel free to remove unnecessory columns and format as per your requirements.


Features:
1. Provides enriched details about the items which are part of groups configured in source, destination and services columns.
2. Detects duplicate rules by comparing source, destination, ports and interfaces.
3. Detects any-any configurations in source, destination and services section.
4. Detects if no security profiles are configured for the rule.
5. Detects if no logging is enabled for the rule.
6. Detects if no traffic is going through the rule (You need to rename traditional fortigate csv export and upload it in the tool)
7. Can enrich the "Previous Client Remarks" from the old report (.xlsx format and report should comtains policy ID and client remarks in same row)

Limitations:
1. Does not supports overlapping rule detections (Please review this manually).
2. Need manual review The observations and recommendations given by the tool if there are any exceptions.

Upcomming Features:
1. Integration with the Database
2. Difference in the current rule as compared to previous configuration
3. Suggestions for security profiles based on the allowed ports
4. Detection of overlapping / shadow rules

Plese write to pandhareamol96@gmail.com to provide a feedback or feature recommendations.
"""

import tkinter as tk
from tkinter import filedialog, IntVar, ttk, messagebox 
import tkinter.constants as tkc

from os import startfile, path
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from ipaddress import ip_address, IPv4Address, IPv4Network

## Constants
APP_NAME = 'Fortigate Firewall Rule Analyzer v1.0'
required_columns = ['rule_id','name','srcaddr','dstaddr','service','action','status','profiles','application-list','captive-portal-exempt','internet-service-name','schedule','logtraffic','uuid','comments','groups','nat','fsso-groups','srcintf','dstintf']

# For excel file
actual_required_columns = ['Rule ID','Rule Name','Src Address','Dst Address','Service','Action','Status','Profiles','Application List','Captive Portal Exempt','Internet Service Name','Schedule','Logtraffic','uuid','Comments','Groups','NAT','fsso-groups','Src Interface','Dst Interface']

enriched_columns = ['rule_id','name','srcaddr','En_srcaddr','srcintf','dstaddr','En_destaddr','dstintf','service','En_services','action','profiles','Additional Info','Severity','Observation','Recommendations','previous_remarks']

# For excel file
actual_enriched_columns = ['Rule ID','Rule Name','Src Address','Enriched Src Address','Src Interface','Dst Address','Enriched Dst Address','Dst Interface','Services / Ports','Enriched Services','Action','Security Profiles','Additional Info','Severity','Observation','Recommendations','Previous Remarks']

column_widths = {
    'A': 7,
    'B': 20,
    'C': 13,
    'D': 20,
    'E': 13,
    'F': 20,
    'G': 13,
    'H': 7,
    'I': 20,
    'J': 31,
    'K': 8,
    'L': 20,
    'M': 20,
    'N': 10,
    'O': 30,
    'P': 30,
    'Q': 30,
}

def read_file(file: str) -> str:
    """Function to read the file and return the content"""
    with open(file,'r') as f:
        lines = f.read()

    return lines


def get_fw_rules(lines: str) -> list:
    """Function to read the file and parse the firewall policy section
    :param file :string (file path)
    :return rules :list (list of lines containing the rules of the firewalls)
    """

    if 'config firewall policy' in lines:
        rules = lines.split('config firewall policy')[1]
        rules = rules.split('\nend')[0]
        lines_ = rules.split('\n')
    else:
        messagebox.showerror(app, "Rules section was not present in the configuration file. Make sure you have selected correct configuration file")
        return None
        

    list_rules = []
    d = {}
    for line in lines_:
        try:
            line = line.strip()
            if 'edit ' in line:
                d.update({'rule_id':line.split()[1]})

            if 'set ' in line:
                split = line.split('set ')[1].strip().split(' ',1)
                if '" "' in split[1]:
                    items = split[1].strip('"').split('" "')
                else:
                    items = split[1].strip('"')

                d.update({split[0]:items})

            if line.strip().lower() == 'next':
                profiles = ''
                additional_info = ''
                list_of_profiles = ['inspection-mode','webfilter-profile','ssl-ssh-profile','av-profile','ips-sensor','dnsfilter-profile','utm-status','users']
                list_additional_fields = ['application-list','captive-portal-exempt','logtraffic','comments','groups','nat','fsso-groups']
                try:
                    for k,v in d.items():
                        if k in list_of_profiles and v not in ['',' ','-']:
                            if isinstance(v,list):
                                v = ','.join(v)
                            profiles += k + ':' + v + '\n'
                        if k in list_additional_fields and v not in ['',' ','-']:
                            if isinstance(v, list):
                                v= ','.join(v)
                            additional_info += k + ':' + v + '\n'
                except Exception as ex:
                    #print(f"Exception: {ex} - {profiles}, {k}, {v}")
                    messagebox.showerror(app, f"There was error while parsing the security profiles and some other fields. Please report this issue to pandhareamol96@gmail.com.\nError: {ex}")

                # Creating a new key for all the policies
                d.update({'profiles':profiles})
                d.update({'Additional Info': additional_info })
                    
                list_rules.append(d)
                d = {}
        except Exception as ex:
            #print(f"Exception: {ex} - {line}")
            messagebox.showerror(app, f"There was error while parsing the security profiles and some other fields. Please report this issue to pandhareamol96@gmail.com.\nError: {ex}\nLine: {line}")
            

    return list_rules


def get_fw_config_section(lines: str, section: str) -> dict:
    """Function to get the contents of firewall address groups"""
    if not lines: return

    if section in lines:
        print(f"Section '{section}' found in config file. Count of sections:- {lines.count(section)}")
        sections = lines.split(section)
    else:
        print(f'{section} section not found in the config file')
        return []

    dict_data = dict()

    for section in sections[1:]:
        rules = section.split('\nend')[0]
        lines_ = rules.split('\n')

        for line in lines_:
            line = line.strip()
            if 'edit ' in line:
                key = ' '.join(line.split()[1:]).strip().strip('"')
                if not key in dict_data:
                    dict_data[key] = {}
                else:
                    print(f"Key '{key}' already available in dictionary.")

            if 'unset' in line:
                continue

            if 'set ' in line:
                
                split = line.split('set ')[1].strip().split(' ',1)
                if '" "' in split[1]:
                    items = split[1].strip('"').split('" "')
                else:
                    items = split[1].strip('"')

                dict_data[key].update({split[0]:items})

            if line.strip().lower() == 'next':
                key = None
                items = None

    return dict_data


def data_list_to_excel(data_list: list,file: str, columns: list, xl_columns: list) -> None:
    """Function to save the report to the excel file"""
    
    wb = Workbook()
    ws = wb.active
    if 'Observation' in columns:
        ws0 = wb.create_sheet('Summary',1)
        ws0.title = 'Summary'
    ws.title = 'Assessment Details'
    ws.sheet_view.zoomScale = 80
    ws.freeze_panes = 'C2'

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    try:
        # Populating the Headers in the Report
        for col_num, header in enumerate(xl_columns, start=1):
            ws.cell(row=1, column=col_num).value = header
            ws.cell(row=1, column=col_num).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws.cell(row=1, column=col_num).font = Font(color="FFFFFF", bold=True)
            #ws.cell(row=1, column=col_num).font = openpyxl.styles.Alignment(wrap_text=True, horizontal="center", vertical="center")

        for row_num, row in enumerate(data_list, start=2):
            for col_num, key in enumerate(columns, start=1):
                value = row.get(key, '')
                if value == '':
                    ws.cell(row=row_num, column=col_num).value = None
                else:
                    if type(value) == list:
                        value = '\n'.join(value)
                    ws.cell(row=row_num, column=col_num).value = value
                    ws.cell(row=row_num, column=col_num).alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
    except Exception as ex:
        messagebox.showerror(app, f"Exception while writing report file.\nError Details: {ex}")

    try:
        wb.save(file)
    except PermissionError:
        messagebox.showwarning(app, f"Permission denied - Close the output file '{file}' if already open.")


def get_member_details(member: str, address_mappings: dict) -> str:
    """
    Function to resolve the member details and return the exact details of the 
    member
    """
    member_details = address_mappings[member]
    if 'type' in member_details:
        if member_details['type'] == 'geography':
            return member
        elif member_details['type'] == 'fqdn':
            return member_details['fqdn']
        elif member_details['type'] == 'iprange':
            return f"{member_details['start-ip']}-{member_details['end-ip']}"
        elif member_details['type'] == 'mac':
            if isinstance(member_details['macaddr'], list):
                return '\n'.join(member_details['macaddr']).strip('\n')
            else:
                return member_details['macaddr']
        elif member_details['type'] == 'interface-subnet':
            return member_details['subnet']
        elif member_details['type'] == 'address':
            return member_details['resource']
        #elif member_details['type'] == 'ipmask':
        #   ip = member_details['subnet'].split(' ')[0].strip()
        #   mask = member_details['subnet'].split(' ')[1].strip()
        #   (cidr, ip_cnt) = calculate_cidr_and_ip_count(ip, mask)
        #   return f"{member_details['subnet']},\nCIDR: {cidr},\n IP_CNT: {ip_cnt}"
        else:
            print(f"Need to handle this condition in code: {member_details}")
            messagebox.showwarning(APP_NAME, f"There is an unhandeled condition in the member enrichment function. Please share this config file with pandhareamol96@gmail.com to fix this issue.\n Member Details: {member_details}")

    elif 'subnet' in member_details:
        if member_details['subnet'].split()[1] == '255.255.255.255':
            return member_details['subnet'].split()[0]
        return member_details['subnet']
    elif 'extip' in member_details:
        return f"{member_details['extip']}\n{member_details['mappedip']}"
    elif 'member' in member_details:
        return member_details['member'] if isinstance(member_details['member'], str) else "\n".join(member_details['member']) 
    else:
        print(f"Need to handle this condition in code: {member_details}")
        messagebox.showwarning(APP_NAME, f"There is an unhandeled condition in the member enrichment function. Please share this config file with pandhareamol96@gmail.com to fix this issue.\n Member Details: {member_details}")


def enrich_helper(addr:str, address_groups:dict, address_mappings:dict, fw_addresses:dict, external_resource:dict, vip:dict) -> list:
    """Function to identify the type of src/dst ips i.e., ip-range, subnet, fqdn,
    geography (country), dynamic, """
    list_data = list()

    if not addr:
        return
    
    if address_groups and addr in address_groups:
        members = address_groups[addr]['member']
        if isinstance(members, list):
            for member in members:
                if member in address_mappings:
                    data = get_member_details(member, address_mappings)
                    list_data.append(data)
                else:
                    #print(f"Group member '{member}' not found in address mappings.")
                    messagebox.showwarning(APP_NAME, f"Member '{member}' not found in the config file. Please share this config file with pandhareamol96@gmail.com to check for possible improvements in data enrichment feature")
                    list_data.append(member)

                    
        elif members in address_mappings:
            data = get_member_details(members, address_mappings)
            list_data.append(data)
        else:
            #print(f"Group member '{members}' not found in the address mappings.")
            messagebox.showwarning(APP_NAME, f"Member '{member}' not found in the config file. Please share this config file with pandhareamol96@gmail.com to check for possible improvements in data enrichment feature")
            list_data.append(member)
               

    elif address_mappings and addr in address_mappings:
        data = get_member_details(addr, address_mappings)
        list_data.append(data)

    elif fw_addresses and addr in fw_addresses:
        data = get_member_details(addr, fw_addresses)
        list_data.append(data)

    elif external_resource and addr in external_resource:
        data = get_member_details(addr, external_resource)
        list_data.append(data)

    elif vip and addr in vip:
        data = get_member_details(addr, vip)
        list_data.append(data)

    return list_data

def is_Valid_ipv4(IP: str) -> str:
    try:
        return True if type(ip_address(IP)) is IPv4Address else False
    except ValueError:
        return False

def calculate_cidr_and_ip_count(ip_address: str, subnet_mask: str) -> tuple[str, int]:
    network = IPv4Network(f"{ip_address}/{subnet_mask}", strict=False)
    cidr = network.with_prefixlen
    ip_count = network.num_addresses
    return cidr, ip_count


def check_ip_type(ip_address: str) -> str:
    # Define private IP address ranges
    private_ranges = [
        IPv4Network('10.0.0.0/8'),
        IPv4Network('172.16.0.0/12'),
        IPv4Network('192.168.0.0/16')
    ]
    
    ip = IPv4Address(ip_address)
    
    # Check if the IP address is in any of the private ranges
    for private_range in private_ranges:
        if ip in private_range:
            return "Internal"
    
    return "External"

def service_enrich_helper(port: str,custom_services:list, service_group:dict) -> list:
    list_ports = list()
    if port in service_group:
        _list_ports = service_group[port]['member']
        for _port in _list_ports:
            if _port in custom_services:
                port_details = custom_services[_port]
                for k,v in port_details.items():
                    list_ports.append(f"{_port} - {k} - {v}")

    else:
        if port in custom_services:
            port_details = custom_services[port]
            for k,v in port_details.items():
                list_ports.append(f"{port} - {k} - {v}")

    return list_ports


def enrich(list_rules,address_mappings=None, address_groups=None, fw_addresses=None, external_resource=None, custom_services=None, service_group=None, vip=None, dict_previous_remarks=None, dict_bandwidth=None):
    """Function to enrich the firewall rules with the actual ip, domain, url values 
    instead of group name. 
    : param list_rules : list (parsed fw rules)
    : param address_mappings : dict (dictionary of the mappings)
    : return enriched_rules
    """
    if not list_rules: return

    if not address_mappings and not address_groups:
        messagebox.showwarning("Can't enrich the rules. Mappings not found.")
        return

    enriched_rules = list()
    disabled_rules = list()
    blocked_rules = list()

    for rule in list_rules:
        if 'status' in rule and rule['status'].lower() == 'disable':
            disabled_rules.append(rule)
            continue

        if not 'action' in rule:
            blocked_rules.append(rule)
            continue

        if not 'action' in rule:
            rule.update({'Enriched Source Address': None})
            rule.update({'Enriched Destination Address': None})
            enriched_rules.append(rule)
            continue

        src_addr = rule['srcaddr'] if 'srcaddr' in rule else ''
        dest_addr = rule['dstaddr'] if 'dstaddr' in rule else ''
        ports = rule['service'] if 'service' in rule else ''

        previous_comments = None
        if dict_previous_remarks and rule['rule_id'] in dict_previous_remarks:
            previous_comments = dict_previous_remarks[rule['rule_id']]

        bytes = None
        additional_info = rule['Additional Info']

        if dict_bandwidth and int(rule['rule_id']) in dict_bandwidth:
            bytes = dict_bandwidth[int(rule['rule_id'])]
            additional_info += 'Bandwidth Usage: ' + bytes

        # Normalize port details
        enriched_ports = list()    
        if isinstance(ports, list):
            for port in ports:
                if port.isnumeric():
                    enriched_ports.append(port)
                else:
                    list_d = service_enrich_helper(port,custom_services, service_group)
                    enriched_ports += list_d
                        
        else:
            if ports.isnumeric():
                enriched_ports.append(ports)
            else:
                list_d = service_enrich_helper(ports,custom_services, service_group)
                enriched_ports += list_d

        
        enriched_src_addr = list()
        if isinstance(src_addr, list):
            for addr in src_addr:
                if not ( is_Valid_ipv4(addr) or addr.lower() == 'all'): # or 'any' in addr.lower()):
                    list_data = enrich_helper(addr, address_groups, address_mappings, fw_addresses, external_resource, vip)
                    enriched_src_addr+= list_data
                else:
                    enriched_src_addr.append(addr)
        else:
            if src_addr and not ( is_Valid_ipv4(src_addr) or src_addr.lower() == 'all'): # or 'any' in src_addr.lower()) or '.' in src_addr:
                list_data = enrich_helper(src_addr, address_groups, address_mappings, fw_addresses, external_resource, vip)
                enriched_src_addr+= list_data
            else:
                enriched_src_addr.append(src_addr)
            
        # Normalize destination address
        enriched_dest_addr = list()
        if isinstance(dest_addr, list):
            for addr in dest_addr:
                if not ( is_Valid_ipv4(addr) or addr.lower() == 'all'): # or 'any' in addr.lower()):
                    # resolve
                    list_data = enrich_helper(addr, address_groups, address_mappings, fw_addresses, external_resource, vip)
                    enriched_dest_addr+= list_data
                else:
                    enriched_dest_addr.append(addr)
        else:
            if dest_addr and not (is_Valid_ipv4(dest_addr) or dest_addr.lower() == 'all'): # or 'any' in dest_addr.lower()):
                # resolve
                list_data = enrich_helper(dest_addr, address_groups, address_mappings, fw_addresses, external_resource, vip)
                enriched_dest_addr+= list_data
            else:
                enriched_dest_addr.append(dest_addr)

        if ( not dest_addr or dest_addr in ['',' ','-',None] ) and (not rule['internet-service-name'] in ['','-',' ', None]):
            enriched_dest_addr = rule['internet-service-name']

        rule.update({'En_srcaddr': enriched_src_addr})
        rule.update({'En_destaddr': enriched_dest_addr})
        rule.update({'En_services': enriched_ports})
        rule.update({'previous_remarks': previous_comments})
        rule.update({'Additional Info': additional_info})
        rule.update({'bytes': bytes})
        enriched_rules.append(rule)

    return enriched_rules, blocked_rules, disabled_rules

def analyze_and_recommend(enriched_rules: list) -> list:
    """
    Function to analyse the enriched rules, add observations and provide recommendations based upon the same. 
    """
    if not enriched_rules: return 

    """
    Sample rule for reference:
    {'rule_id': '223', 'name': 'RND LAB', 'uuid': '29279c14-ee62-51ed-b63a-f5bcf42df591', 'srcintf': ['LAN', 'port13'], 'dstintf': 'upg-zone-port3', 'action': 'accept', 'srcaddr': ['1.1.1.1', '2.2.2.2'], 'dstaddr': 'all', 'schedule': 'always', 'service': 'ALL', 'utm-status': 'enable', 'inspection-mode': 'proxy', 'ssl-ssh-profile': 'certificate-inspection', 'av-profile': 'XYZ AV Filter', 'webfilter-profile': 'XYZ Web Filter', 'ips-sensor': 'XYZ IPS Filter', 'logtraffic': 'all', 'nat': 'enable', 'traffic-shaper': 'RND LAB', 'profiles': 'utm-status:enable\ninspection-mode:proxy\nssl-ssh-profile:certificate-inspection\nav-profile:XYZ AV Filter\nwebfilter-profile:XYZ Web Filter\nips-sensor:CEAT-Halol IPS Filter\n', 'En_srcaddr': ['1.1.1.1', '2.2.2.2'], 'En_destaddr': ['all'], 'En_services': ['ALL - category - General', 'ALL - protocol - IP']}
    """

    severity_score = 0
    dict_severity_score = { 0: 'None',
                            1: 'Low', 
                            2: 'Medium',
                            3: 'High',
                            4: 'Critical'}

    print(f"Enriched Rules: {enriched_rules}")
    for rule in enriched_rules:
        severity = None
        observations = ''
        recommendations = ''
        obs_cnt = 1
        rcm_cnt = 1
        severity_score = 0
        
        ## Any Source is allowed
        try:
            if 'En_srcaddr' in rule and 'all' in [x.lower() for x in rule['En_srcaddr']]:
                observations = observations + '{0}. Any source is allowed.\n'.format(obs_cnt)
                recommendations = recommendations + '{0}. Allow only required source ips / domains.\n'.format(rcm_cnt)
                obs_cnt += 1
                rcm_cnt += 1
                severity_score += 1
        except Exception as ex:
            print(f"Exception while analyzing srcaddr:  {ex} - data:  {rule['En_srcaddr']}")
            

        ## Any destination is allowed
        try:
            if 'En_destaddr' in rule and 'all' in [x.lower() for x in rule['En_destaddr']] and not 'captive-portal-exempt' in rule:
                observations = observations + '{0}. Any destination is allowed.\n'.format(obs_cnt)
                recommendations = recommendations + '{0}. Allow only required destination ips / domains.\n'.format(rcm_cnt)
                obs_cnt += 1
                rcm_cnt += 1
                severity_score += 1
        except Exception as ex:
            print(f"Exception while analyzing srcaddr:  {ex} - data:  {rule['En_srcaddr']}")

        ## Any Services are allowed
        if 'En_services' in rule:
            list_ =  [x.lower() for x in rule['En_services'] if 'all ' in x.lower()]
            if rule['En_services'] == 'all' or len(list_) > 0:
                observations = observations + '{0}. Any service is allowed.\n'.format(obs_cnt)
                recommendations = recommendations + '{0}. Allow only required services.\n'.format(rcm_cnt)
                obs_cnt += 1
                rcm_cnt += 1
                severity_score += 1

        ## No Security profile is configured
        if 'profiles' in rule and rule['profiles'] in ['',' ','-',None]:
            observations = observations + '{0}. There is no security profile enabled for the rule.\n'.format(obs_cnt)
            recommendations = recommendations + '{0}. Enable required security profiles. (AV, UTM, SSL)\n'.format(rcm_cnt)
            obs_cnt += 1
            rcm_cnt += 1
            severity_score += 1

        ## Rule name is not defined
        if not 'name' in rule:
            observations = observations + '{0}. Rule name is not defined.\n'.format(obs_cnt)
            recommendations = recommendations + '{0}. Give an explanatory name to the rule.\n'.format(rcm_cnt)
            obs_cnt += 1
            rcm_cnt += 1

        ## Logging is not enabled
        if 'logtraffic' in rule and rule['logtraffic'].lower() == 'disabled':
            observations = observations + '{0}. Logging is not enabled for this rule.\n'.format(obs_cnt)
            recommendations = recommendations + '{0}. Enable logging / log forwarding for this rule.\n'.format(rcm_cnt)
            obs_cnt += 1
            rcm_cnt += 1

        ## No traffic is going through the rule
        if 'bytes' in rule and rule['bytes'] == '0 B':
            observations = observations + '{0}. There is no traffic going through this policy.\n'.format(obs_cnt)
            recommendations = recommendations + '{0}. Disable this policy if not required for the business purpose..\n'.format(rcm_cnt)
            obs_cnt += 1
            rcm_cnt += 1


        # Validating the dupilcate rules based on src, dst, interface and ports.
        duplicate_cnt = 0   # Checks how many parameters are duplicate
        duplicate_rules = []
        for ru in enriched_rules:
            if rule['rule_id'] == ru['rule_id']:
                continue

            try:
                if not'En_srcaddr' in rule and not 'En_srcaddr' in ru:
                    duplicate_cnt += 1
                elif rule['En_srcaddr'] == ru['En_srcaddr']:
                    duplicate_cnt += 1
            except KeyError:
                pass

            try:
                if not 'En_destaddr' in rule and not 'En_destaddr' in ru:
                    duplicate_cnt += 1
                elif rule['En_destaddr'] == ru['En_destaddr']:
                    duplicate_cnt += 1
            except KeyError:
                pass

            try:
                if not 'En_services' in rule and not 'En_services' in ru:
                    duplicate_cnt += 1
                elif rule['En_services'] == ru['En_services']:
                    duplicate_cnt += 1
            except KeyError:
                pass

            try:
                if not 'srcintf' in rule and not 'srcintf' in ru:
                    duplicate_cnt += 1
                elif rule['srcintf'] == ru['srcintf']:
                    duplicate_cnt += 1
            except KeyError:
                pass

            try:
                if not 'dstintf' in rule and not 'dstintf' in ru:
                    duplicate_cnt += 1
                elif rule['dstintf'] == ru['dstintf']:
                    duplicate_cnt += 1
            except KeyError:
                pass

            if duplicate_cnt == 5:
                duplicate_rules.append(ru['rule_id'])
                
            duplicate_cnt = 0

        if len(duplicate_rules) > 0:
            if len(duplicate_rules) > 1:
                observations = observations + "{0}. Rules {1} are duplicates of this rule.\n".format(obs_cnt, ', '.join(duplicate_rules))
                recommendations = recommendations + '{0}. You can disable / delete the duplicate rules and keep any one of them.\n'.format(rcm_cnt)
                obs_cnt += 1
                rcm_cnt += 1
            else:
                observations = observations + "{0}. Rule {1} is duplicate of this rule.\n".format(obs_cnt, ', '.join(duplicate_rules))
                recommendations = recommendations + '{0}. You can disable / delete the duplicate rules and keep any one of them.\n'.format(rcm_cnt)
                obs_cnt += 1
                rcm_cnt += 1

        # Adjusting the score to the currently expected value. Change it later
        if severity_score > 4:
            severity_score = 4

        # Assigning the precalculated severity
        severity = dict_severity_score[severity_score]

        rule.update({'Observation': observations})
        rule.update({'Recommendations': recommendations})
        rule.update({'Severity': severity})

    return enriched_rules



def create_dict_from_xlsx(file_path:str, sheet_name:str, key_column:str, value_column:str) -> dict:
    """
    Read an XLSX file and create a dictionary using the specified key and value columns.
    Returns:
        dict: A dictionary created from the given columns.
    """
    try:
        workbook = load_workbook(file_path)

        sheet = workbook[sheet_name]
        
        # Find the column indices for the specified key and value columns
        key_index, value_index = None, None
        for col_index, cell in enumerate(sheet[1], 1):
            if cell.value == key_column:
                key_index = col_index
            if cell.value == value_column:
                value_index = col_index
        
        if key_index is None or value_index is None:
            raise ValueError("Key or value column not found in the file.")
        
        data_dict = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = row[key_index - 1]  # Subtract 1 to convert to 0-based index
            value = row[value_index - 1]
            data_dict[key] = value
        
        return data_dict
    except Exception:
        messagebox.showerror(APP_NAME, "Error while enriching data using the uploaded file. Please verify the file content or report the issue to pandhareamol96@gmail.com")
        return None


# GUI Part
    
def get_bandwidth_file():
    app.bandwidth_report = filedialog.askopenfilename(initialdir='\\', title='Select report file.', filetypes=[("Excel File",".xlsx")])
    
    bandwidth_report_filename_label.configure(text=app.bandwidth_report.split('/',1)[-1])

    # Updating the fields in the previous report into the user inputs
    workbook = load_workbook(app.bandwidth_report)

    # Populating the sheet Name
    b_sheet_names = workbook.sheetnames
    workbook.close()
    print(f"Sheet Names: {b_sheet_names}")
    for sheet in b_sheet_names:
        b_sheet_listbox.insert(0, sheet)


def get_previous_report_file():
    app.previous_report = filedialog.askopenfilename(initialdir='\\', title='Select report file.', filetypes=[("Report File",".xlsx")])
    
    #ttk.Label(app, text=app.previous_report.rsplit('/',1)[-1]).place(x=20, y=245)
    previous_report_filename_label.configure(text=app.previous_report.rsplit('/',1)[-1])

    # Updating the fields in the previous report into the user inputs
    workbook = load_workbook(app.previous_report)

    # Populating the sheet Name
    sheet_names = workbook.sheetnames
    workbook.close()
    print(f"Sheet Names: {sheet_names}")
    for sheet in sheet_names:
        sheet_listbox.insert(0, sheet)


    
def confirm_sheet_name_and_populate_columns():
    # Populating the rule ID and Client Remark columns
    workbook = load_workbook(app.previous_report)
    print("Inside confirm sheet name and polulate columns")
    sheet = workbook[app.old_sheet_name]

    columns = sheet[1]
    
    for column in columns:
        rule_id_listbox.insert(0, column.value)
        remarks_listbox.insert(0, column.value)

    workbook.close()

def confirm_b_sheet_name_and_populate_columns():
    # Populating the rule ID and Client Remark columns
    b_workbook = load_workbook(app.bandwidth_report)
    print("Inside confirm bandwidth sheet name and polulate columns")
    sheet = b_workbook[app.bandwidth_sheet_name]

    columns = sheet[1]
    
    for column in columns:
        b_rule_id_listbox.insert(0, column.value)
        b_remarks_listbox.insert(0, column.value)

    b_workbook.close()


def get_firewall_config_file():
    app.config_filename = filedialog.askopenfilename(initialdir='\\', title='Select Firewall confif file', filetypes=[("Config Files",".conf")])
    config_filename_label.configure(text=app.config_filename.rsplit('/', 1)[-1])
    set_state_of_submit_button()


def set_output_directory():
    app.output_foldername = filedialog.askdirectory()
    output_directory_label.configure(text=app.output_foldername.rsplit('/',1)[-1])
    set_state_of_submit_button()


def set_filename():
    """
    This function will be executed when user confirms the output filename.
    """
    try:
        app._filename = f'{name_var.get()}.xlsx'
        inputtxt.config(state=tkc.DISABLED)
        set_state_of_submit_button()
    except Exception as ex:
        messagebox.showerror("Error", f"Exception while setting filename: {ex}")


def get_output_filepath():
    """
    This function will be called to prepare complete output filepath before preparing the report.
    """
    try:
        return app.output_foldername + '\\' + app._filename 
    except Exception:
        messagebox.showerror("Error", "Output filename was not given")


def set_sheetname(event):
    app.old_sheet_name = sheet_listbox.get(tkc.ACTIVE)
    sheet_label.config(text=f"Sheet Selected: {app.old_sheet_name}")


def set_b_sheetname(event):
    app.bandwidth_sheet_name = b_sheet_listbox.get(tkc.ACTIVE)
    b_sheet_label.config(text=f"Sheet Selected: {app.bandwidth_sheet_name}")


def set_rule_id_column(event):
    app.rule_id_for_old_comments = rule_id_listbox.get(tkc.ACTIVE)
    rule_id_label.config(text=f'Selected: {app.rule_id_for_old_comments}')

def set_b_rule_id_column(event):
    app.rule_id_for_bandwidth = b_rule_id_listbox.get(tkc.ACTIVE)
    b_rule_id_label.config(text=f'Selected: {app.rule_id_for_bandwidth}')


def set_remarks_column(event):
    app.remarks_column = remarks_listbox.get(tkc.ACTIVE)
    remarks_label.config(text=f'Selected: {app.remarks_column}')

def set_b_remarks_column(event):
    app.bandwidth_column = b_remarks_listbox.get(tkc.ACTIVE)
    b_remarks_label.config(text=f'Selected: {app.bandwidth_column}')

def enable_options_for_previous_report():
    """
    Enabling the Buttons, Text boxes and Labels when enrich previous comments checkbox is selected
    """
    submit_button.place(x=300, y=600)
    if previous_comments.get():
        app.geometry("900x700")
        previous_report.place(x=20, y=220)
        sheet_label.place(x=20, y=280)
        sheet_listbox.place(x=20, y=300)
        confirm_sheet.place(x=20, y=400)
        rule_id_label.place(x=20, y=450)
        rule_id_listbox.place(x=20, y=480)
        remarks_label.place(x=200, y=450)
        remarks_listbox.place(x=200, y=480)
        previous_report_filename_label.place(x=20, y=245)
    else:
        if not bandwidth_details.get():
            app.geometry("700x400")
            submit_button.place(x=300, y=230)
        previous_report.place_forget()
        sheet_label.place_forget()
        sheet_listbox.place_forget()
        confirm_sheet.place_forget()
        rule_id_label.place_forget()
        rule_id_listbox.place_forget()
        remarks_label.place_forget()
        remarks_listbox.place_forget()
        previous_report_filename_label.place_forget()

        # clearing all the veriables for previus report
        clear_previous_report_variables()

def clear_previous_report_variables():
    previous_report_filename_label.configure(text='')
    sheet_label.configure(text=None)
    rule_id_label.configure(text=None)
    remarks_label.configure(text=None)
    sheet_name.set(None)
    sheet_listbox.delete(0, tkc.END)
    rule_id_listbox.delete(0, tkc.END)
    remarks_listbox.delete(0, tkc.END)


def enable_options_for_bandwidth_report():
    """
    Enabling the Buttons, Text boxes and Labels when enrich bandwidth checkbox is selected
    """
    
    if bandwidth_details.get():
        app.geometry("900x700")
        submit_button.place(x=300, y=600)
        bandwidth_report.place(x=400, y=220)
        b_sheet_label.place(x=400, y=280)
        b_sheet_listbox.place(x=400, y=300)
        b_confirm_sheet.place(x=400, y=400)
        b_rule_id_label.place(x=400, y=450)
        b_rule_id_listbox.place(x=400, y=480)
        b_remarks_label.place(x=585, y=450)
        b_remarks_listbox.place(x=585, y=480)
        bandwidth_report_filename_label.place(x=400, y=245)
    else:
        if not previous_comments.get():
            app.geometry("700x400")
            submit_button.place(x=300, y=230)
        bandwidth_report.place_forget()
        b_sheet_label.place_forget()
        b_sheet_listbox.place_forget()
        b_confirm_sheet.place_forget()
        b_rule_id_label.place_forget()
        b_rule_id_listbox.place_forget()
        b_remarks_label.place_forget()
        b_remarks_listbox.place_forget()
        bandwidth_report_filename_label.place_forget()

        # Clearing bandwidth report variables
        clear_bandwidth_report_variables()


def clear_bandwidth_report_variables():
    bandwidth_report_filename_label.configure(text='')
    b_sheet_label.configure(text=None)
    b_rule_id_label.configure(text=None)
    b_remarks_label.configure(text=None)
    bandwidth_sheet_name.set(None)
    b_sheet_listbox.delete(0, tkc.END)
    b_rule_id_listbox.delete(0, tkc.END)
    b_remarks_listbox.delete(0, tkc.END)


def validate_user_input(text):
    if text.isalnum() or ('_' in text and text.replace('_','').isalnum()):
        filename_confirm_button.configure(state=tkc.ACTIVE)
        return True
    else:
        
        return False


def process_data(config_file, enrich_, report_, previous_comments, bandwidth_details):
    # Reading the raw lines from the config file
    raw_lines =  read_file(config_file)

    # Parsing the firewall rules.
    list_rules = get_fw_rules(raw_lines)
    if not list_rules:
        return

    enriched_rules = None
    if enrich_:
        # Parsing the config firewall address mappings. 
        address_mappings = get_fw_config_section(raw_lines, 'config firewall address')

        # Parsing the config firewall address groups
        address_groups = get_fw_config_section(raw_lines, 'config firewall addrgrp')

        # Parsing firewall addresses
        fw_addresses = None
        external_resource = get_fw_config_section(raw_lines, 'system external-resource')

        # Parsing firewall services
        custom_services = None
        service_group = None
        #custom_services = get_services_mapping(raw_lines, 'config firewall service custom')
        custom_services = get_fw_config_section(raw_lines, 'config firewall service custom')
        service_group = get_fw_config_section(raw_lines, 'config firewall service group')

        # Parsing the vip section
        vip = get_fw_config_section(raw_lines, 'config firewall vip')  # not used in the enrich function yet

        dict_previous_comments = None
        if previous_comments:
            dict_previous_comments = create_dict_from_xlsx(app.previous_report, app.old_sheet_name, app.rule_id_for_old_comments, app.remarks_column)

        dict_bandwidth = None
        print(f"Selected bandwidth Details: {bandwidth_details}")
        if bandwidth_details:
            dict_bandwidth = create_dict_from_xlsx(app.bandwidth_report,app.bandwidth_sheet_name,app.rule_id_for_bandwidth,app.bandwidth_column)


        # Enrich the rules using the address mappings.
        (enriched_rules, blocked_rules, disabled_rules) = enrich(list_rules, address_mappings, address_groups, fw_addresses, external_resource, custom_services, service_group, vip, dict_previous_comments, dict_bandwidth)


    analyzed_rules = None
    if enriched_rules and report_:
        analyzed_rules = analyze_and_recommend(enriched_rules)


    ## Save the output in the xlsx file
    output_filepath = get_output_filepath()

    if output_filepath:
        if enrich_ and enriched_rules:
            print('Number of rules after enrichment: ',len(enriched_rules))

            if report_ and analyzed_rules:
                data_list_to_excel(analyzed_rules, output_filepath, enriched_columns, actual_enriched_columns)
                
            else:
                data_list_to_excel(enriched_rules, output_filepath, enriched_columns, actual_enriched_columns)
        else:
            data_list_to_excel(list_rules, output_filepath, required_columns, actual_required_columns)

    messagebox.showinfo(APP_NAME,
                        f"""
                        Total Rules: {len(enriched_rules) + len(disabled_rules) + len(blocked_rules)}
                        Blocked Rules: {len(blocked_rules)}
                        Disabled_Rules: {len(disabled_rules)}
                        """)
    startfile(output_filepath)

       


def reset_selection():
    """
    Function to clear all the selections and text boxes
    """
    config_filename_label.configure(text='')
    output_directory_label.configure(text='')
    inputtxt.config(state=tkc.ACTIVE)
    filename_confirm_button.configure(state=tkc.DISABLED)
    name_var.set('')
    enrich_.set(0)
    report_.set(0)
    previous_comments.set(0)
    bandwidth_details.set(0)
    submit_button.configure(state=tkc.DISABLED)

    enable_options_for_bandwidth_report()
    enable_options_for_previous_report()


def set_state_of_submit_button():
    """
    Function to check all the mandatory inputs and assign the state to the submit button.
    This will reduce the unexpected behaviour of the application by controlling the user input.
    """
    try:
        if (    app.config_filename not in ['',' ', None] and 
                app.output_foldername not in ['', ' ', None] and 
                app._filename not in ['', ' ', None] and
                ( enrich_.get() or report_.get() ) ):
            
            submit_button.configure(state=tkc.ACTIVE)
        else:
            submit_button.configure(state=tkc.DISABLED)
    except AttributeError:
        pass


def show_info():
    """
    Function to show the documentation of the application to the user.
    """
    messagebox.showinfo(APP_NAME, tool_info)


def set_logo(app):
    """
    Sets a logo (icon) for a Tkinter application window.

    Parameters:
    - app: The main Tkinter window (Tk object).
    - logo_path: Path to the icon file (.ico for Windows, .png might work on some systems).
    """
    try:
        if path.isfile('logo.ico'):
            app.iconbitmap('.\\logo.ico')
        elif path.isfile('.\\logo.png'):
            logo =tk.PhotoImage(file='.\\logo.png')
            app.iconphoto(True, logo)
    except Exception as ex:
        logging.warning(f"Error while configuring the logo - {ex}")


if __name__ == '__main__':
    app = tk.Tk()
    app.title(APP_NAME)
    set_logo(app)
    app.geometry("700x400")

    logs = list()
    app.old_sheet_name = None
    app.rule_id_for_old_comments = None
    app.remarks_column = None

    text_checker =  app.register(validate_user_input)

    reset_button = ttk.Button(app, text= "Reset Selection", command=lambda: reset_selection())
    reset_button.place(x=15, y=5)
    info_button = ttk.Button(app, text='About', command=lambda: show_info())
    info_button.place(relx=1.0, rely=0.0, anchor=tk.NE)


    config_file = ttk.Button(app, text='Select Config File', command=get_firewall_config_file).place(x=15,y=35)
    config_filename_label = ttk.Label(app, text='')
    config_filename_label.place(x=170, y=35, width=500)

    output_directory = ttk.Button(app, text='Select Output Directory', command=set_output_directory).place(x=15, y= 65)
    output_directory_label = ttk.Label(app, text='')
    output_directory_label.place(x=170, y=55, width=500)

    #Capture output filename
    ttk.Label(app, text="Type Output Filename And Click On Confirm Button.").place(x=15, y=100)
    name_var=tk.StringVar()
    name_var.set("")
    inputtxt = ttk.Entry(app, textvariable = name_var, width = 30 , validatecommand=(text_checker, "%P"), validate="key")

    inputtxt.place(x=15, y= 120)
    ttk.Label(app,text='.xlsx').place(x=200, y=125)
    filename_confirm_button =  ttk.Button(app, text = "Confirm", state=tkc.DISABLED,  command=lambda: set_filename())
    filename_confirm_button.place(x=250, y=120)



    enrich_ = IntVar()
    report_ = IntVar()
    previous_comments = IntVar()
    bandwidth_details = IntVar()

    Button1 = ttk.Checkbutton(app, text = "Enrich",  
                        variable = enrich_, 
                        onvalue  = 1, 
                        offvalue = 0,
                        command  = set_state_of_submit_button).place(x=20, y=160)

    Button2 = ttk.Checkbutton(app, text = "Analysis", 
                        variable = report_, 
                        onvalue  = 1, 
                        offvalue = 0,
                        command  = set_state_of_submit_button).place(x=150, y=160)

    Button3 = ttk.Checkbutton(app, text = "Enrich Previous Comments (.xlsx File)", 
                        variable = previous_comments, 
                        onvalue  = 1, 
                        offvalue = 0,
                        command  = enable_options_for_previous_report).place(x=20, y=190)

    Button4 = ttk.Checkbutton(app, text = "Enrich Bandwidth Details (.xlsx File)", 
                        variable = bandwidth_details, 
                        onvalue  = 1, 
                        offvalue = 0,
                        command  = enable_options_for_bandwidth_report).place(x=400, y=190)

    """
    Selecting the previous report to enrich the last client remarks for reported Policy IDs.
    """
    previous_report = ttk.Button(app, text='Select Previous Report File To Enrich Client Remarks.', command=get_previous_report_file)
    previous_report_filename_label = ttk.Label(app, text='')

    sheet_label   = ttk.Label(app, text='Select Sheet Name From The Below Options')
    sheet_name    = tk.StringVar()
    sheet_listbox = tk.Listbox(app, height=5, width=25)
    sheet_listbox.bind("<ButtonRelease-1>", set_sheetname)

    confirm_sheet = ttk.Button(app, text='Confirm Sheet Name', command= confirm_sheet_name_and_populate_columns)


    rule_id_label    = ttk.Label(app, text='Select Rule ID column')
    rule_id_column   = tk.StringVar()
    rule_id_listbox  = tk.Listbox(app,height=5, width=25)
    rule_id_listbox.bind("<ButtonRelease-1>", set_rule_id_column)

    remarks_label   = ttk.Label(app, text='Select Client Remarks Column')
    remarks_column  = tk.StringVar()
    remarks_listbox = tk.Listbox(app,height=5, width=25)
    remarks_listbox.bind("<ButtonRelease-1>", set_remarks_column)


    """
    Enrich the details of data transfered via the ports
    """
    bandwidth_report = ttk.Button(app, 
                                  text='Select file Having Bandwidth Consumption Details', 
                                  command=get_bandwidth_file)
    bandwidth_report_filename_label = ttk.Label(app, text='')

    b_sheet_label        = ttk.Label(app, text='Select Sheet Name From The Below Options')
    bandwidth_sheet_name = tk.StringVar()
    b_sheet_listbox      = tk.Listbox(app, height=5, width=25)
    b_sheet_listbox.bind("<ButtonRelease-1>", set_b_sheetname)

    b_confirm_sheet = ttk.Button(app, 
                                 text='Confirm Sheet Name', 
                                 command= confirm_b_sheet_name_and_populate_columns)


    b_rule_id_label   = ttk.Label(app, text='Select Rule ID column')
    b_rule_id_column  = tk.StringVar()
    b_rule_id_listbox = tk.Listbox(app,height=5, width=25)
    b_rule_id_listbox.bind("<ButtonRelease-1>", set_b_rule_id_column)

    b_remarks_label   = ttk.Label(app, text='Select Bandwidth Column')
    b_remarks_column  = tk.StringVar()
    b_remarks_listbox = tk.Listbox(app,height=5, width=25)
    b_remarks_listbox.bind("<ButtonRelease-1>", set_b_remarks_column)

    ## Generating the report
    try:
        submit_button = ttk.Button(app, 
                                   text= "Generate Report",
                                   state=tkc.DISABLED, 
                                   command=lambda: process_data(app.config_filename, enrich_.get(), report_.get(), previous_comments.get(), bandwidth_details.get()))
    except Exception as e:
        messagebox.showerror("Error", e)
    submit_button.place(x=300, y=230)

    app.mainloop()